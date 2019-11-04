import os
import openpyxl

class ExcelOperations:
    workBookLoc = os.getcwd()+"\\ExcelData.xlsx"
    workbook = openpyxl.load_workbook(workBookLoc)

    def readExcelData(self, sheetname, cell):
        worksheet = ExcelOperations.workbook[sheetname]
        cellVal = worksheet[cell].value
        return cellVal

    def writeExcelData(self, sheetname, cell, data):
        worksheet = ExcelOperations.workbook[sheetname]
        worksheet[cell] = data
        ExcelOperations.workbook.save(ExcelOperations.workBookLoc)

    def maxRowData(self, sheetname):
        worksheet = ExcelOperations.workbook[sheetname]
        return worksheet.max_row

    def maxcolnData(self, sheetname):
        worksheet = ExcelOperations.workbook[sheetname]
        return worksheet.max_column


obj = ExcelOperations()
# for i in range(1, 11):
#     cellValue = "B"+str(i)
#     print(obj.readExcelData("Chapter1_Data", cellValue))
#
# for i in range(2,10):
#     cellValue = "D"+str(i)
#     print(obj.writeExcelData("Chapter1_Data", cellValue, "FAIL"))
#
# obj.writeExcelData("Chapter1_Data", 'D10', "FAIL")
print(obj.readExcelData("Chapter1_Data", "B3"))
print(obj.readExcelData("Chapter1_Data", "D4"))
print(obj.maxRowData("Chapter1_Data"))
print(obj.maxcolnData("Chapter1_Data"))